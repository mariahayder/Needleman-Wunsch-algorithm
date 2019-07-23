#last modification: 23.07, Marysia Hayder

from NeedlemanWunsch import createMatrix, fillMatrix, readMatrix, countMatrix, printMatrix
import os, sys
from openpyxl import Workbook, load_workbook

EXPERIMENT_SET = "diff=1"

def saveExperiment(seqA, seqB, alignedA, alignedB, percentage, fileName):
    list = [" Sequence 1: ", seqA, "\n", " Sequence 2: ", seqB, "\n", " Percentage (alignment-derived): ",
            str(percentage), "\n", "Alignment: ", "\n",
            str(alignedA), "\n", str(alignedB)]
    os.chdir("diff=1")
    file = open(fileName, "w")
    file.writelines(list)
    file.close()



def work(fileA, fileB, fileName):
    file=open(fileA, "r")
    seqA=file.readline()
    file.close()
    file = open(fileB, "r")
    seqB = file.readline()
    file.close()
    matrix = createMatrix(seqA, seqB)
    matrix = fillMatrix(matrix, seqA, seqB)
    alignedA, alignedB = readMatrix(matrix, seqA, seqB)
    percentage = countMatrix(alignedA, alignedB)
    saveExperiment(seqA, seqB, alignedA, alignedB, percentage, fileName)
    return percentage


def createExcelFile(comparedLetters):
    wb = Workbook()
    ws = wb.active
    for i in range(1, len(comparedLetters) + 1):
        d = ws.cell(row=2, column=i+2, value=comparedLetters[i-1])
    for i in range(2, len(comparedLetters) + 1):
        d = ws.cell(row = i + 1, column=2, value=comparedLetters[i - 1])
    wb.save('diff=1.xlsx')

def main():
    comparedLetters = ["n", "i", "c", "e", "f", "g", "h", "l", "s", "t", "u", "y", "polish l", "modificated e"]
    pathname = os.path.dirname(sys.argv[0])
    fullPath = os.path.abspath(pathname)
    os.mkdir(EXPERIMENT_SET)
    createExcelFile(comparedLetters)
    wb = load_workbook('diff=1.xlsx')
    ws = wb.active
    for i in range(len(comparedLetters)):
        for j in range(i + 1, len(comparedLetters)):
            pathname = os.path.dirname(sys.argv[0])
            fullPath=os.path.abspath(pathname)
            os.chdir(fullPath)
            fileName = []
            fileName.append(comparedLetters[i])
            fileName.append('_')
            fileName.append(comparedLetters[j])
            fileName = str(fileName)
            percentage = work(comparedLetters[i], comparedLetters[j], fileName)
            d = ws.cell(row=j+3, column=i + 3, value=percentage)
            wb.save('diff=1.xlsx')
    return 0


if __name__ == '__main__':
    import sys

    sys.exit(main())